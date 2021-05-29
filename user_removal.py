# Removal of User from Github team which is part of an org

import sys
import requests
from requests.auth import HTTPBasicAuth
from openpyxl import Workbook, load_workbook
from pprint import pprint
from zipfile import BadZipFile


def validate_user(membername):
    url = f'{github_url}/users/{membername}'
    try:
        user_data = requests.get(url, auth = HTTPBasicAuth(username, passwd)).json()
    except:
        print(f'ERROR: error fetching information for user {membername}')
        return False
    else:
        if 'message' in user_data.keys() and user_data['message'] == 'Not Found':
            return False
        return True


def get_mem_count(orgname, teamname):
    url = f'{github_url}/orgs/{orgname}/teams/{teamname}'
    try:
        user_data = requests.get(url, auth = HTTPBasicAuth(username, passwd)).json()
    except:
        print(f'ERROR: error fetching information for team {teamname}')
        return None
    else:
        if 'members_count' in user_data.keys():
            mem_count = dict(user_data)['members_count']
            return mem_count
        return None


def list_members(orgname, teamname):
    url = f'{github_url}/orgs/{orgname}/teams/{teamname}/members'
    try:
        user_data = requests.get(url, auth = HTTPBasicAuth(username, passwd))
    except:
        print(f'ERROR: error fetching members information for team {teamname}')
        return None
    else:
        team_mems = [x['login'] for x in list(user_data.json())]
        return team_mems


def add_member(orgname, teamname, membername):
    url = f'{github_url}/orgs/{orgname}/teams/{teamname}/memberships/{membername}'
    try:
        user_data = requests.put(url, auth = HTTPBasicAuth(username, passwd))
    except:
        print(f'ERROR: error while adding member {membername} to team {teamname}')
        return None
    else:
        return user_data.status_code


def remove_member(orgname, teamname, membername):
    url = f'{github_url}/orgs/{orgname}/teams/{teamname}/memberships/{membername}'
    try:
        user_data = requests.delete(url, auth = HTTPBasicAuth(username, passwd))
    except:
        print(f'ERROR: error while removing member {membername} from team {teamname}')
        return None
    else:
        return user_data.status_code


def fetch_data(data_file):
    wb = None
    xl_list=[]
    print('INFO:  Loading data from Excel file..')
    try:
        wb = load_workbook(data_file)
    except FileNotFoundError:
        print(f'ERROR: error loading {data_file}, file not found')
        exit()
    except BadZipFile:
        print(f'ERROR: error loading {data_file}, Not a valid excel file')
        exit()
    except:
        print(f'ERROR: error loading {data_file}, {sys.exc_info()[1]}')
        exit()
    else:
        print('INFO:  Excel workbook loaded successfully')
    ws = wb.active
    row_count = ws.max_row
    print(f'INFO:  Number of rows to process: {row_count}')
    if row_count > 1:
        for row in range(2,row_count+1):
            temp_list=[]
            temp_list.append(ws.cell(row,1).value)
            temp_list.append(ws.cell(row,2).value)
            xl_list.append(temp_list)
    wb.close()
    return xl_list


def process_list(xl_list):
    orgname = teamname = membername = ''
    team_mems = []
    curr_record = 2
    for item_list in xl_list:
        print(f'INFO:  Processing row no. {curr_record}: {item_list}')
        curr_record += 1
        membername = item_list[0]
        orgname = item_list[1].split('_')[0]
        teamname = item_list[1].split('_')[2]
        if not validate_user(membername):
            print(f'ERROR: Username {membername} could not be found')
            continue
        mem_count = get_mem_count(orgname, teamname)
        if mem_count != None and mem_count > 0:
            print(f'INFO:  Members count in team: {mem_count}')
        else:
            print('ERROR: Organization or Team information is incorrect')
            continue
        team_mems = list_members(orgname, teamname)
        if team_mems != None and membername in team_mems:
            print(f'INFO:  Team member {membername} is available in team {teamname}. Removing now..')
            result = remove_member(orgname, teamname, membername)
            if result == 204:
                print(f'INFO:  Removed user {membername} successfully')
            else:
                print(f'ERROR:  User {membername} could not be removed !')
        else:
            print(f'INFO:  Team member {membername} not in the team {teamname}. Adding now..')
            result = add_member(orgname, teamname, membername)
            if result == 200:
                print(f'INFO:  Added user {membername} successfully')
            else:
                print(f'ERROR: User {membername} could not be added !')


def usage():
    usage_str = '''
Usage: python user_removal.py <excel_file>
<excel_file> must be a valid excel file with extension .xlsx
'''
    print(usage_str)


def main():
    arg_count = len(sys.argv)
    mem_dict = None
    if arg_count != 2:
        print('ERROR: Incorrect number of arguments !')
        usage()
        exit()
    data_file = sys.argv[1]
    if data_file[-5:] != '.xlsx':
        print('ERROR: Incorrect file argument !')
        usage()
        exit()
    xl_list = fetch_data(data_file)
    if len(xl_list) != 0:
        process_list(xl_list)


username = '****'
passwd = '****'
github_url = 'https://api.github.com'

if __name__ == '__main__':
    main()
